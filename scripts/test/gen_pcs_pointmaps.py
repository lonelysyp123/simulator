#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 TRINA MV-EMS Modbus TCP 协议 V0.15 Excel 生成 10MW / 5.5MW 两份 EMU 点表。

⚠ 警告: 本脚本全量重写目标 CSV，会覆盖任何手工修改（含点位增删、绑定调整、
描述微调），且点位删除会导致 yc 序号重编号。点表 CSV 是运行期源文件，允许
手工维护；本脚本仅用于从协议 Excel 重新引导（bootstrap）点表。日常增删点位
请直接编辑 CSV，不要重跑本脚本。

数据来源:
- PCS System(PCS 系统) 工作表: 各型号 PCS 单元与 PCS 组的读寄存器(04)和保持寄存器(06/16)
- PCS Status&Alarm(PCS状态和报警) 工作表: 位级定义, 并入对应故障/状态字寄存器描述
- 现有全量 emu.csv: 按 (功能码, 地址) 继承 ModelSim 绑定
"""
import csv
import re
import sys
import openpyxl

ROOT = '/Users/songyinpei/jobs/trina/EssSimulator'
XLSX = f'{ROOT}/TRINA MV-EMS Modbus TCP protocol(天合中压系统对外Modbus TCP通信协议)_V0.15_2026.08.13(1).xlsx'
OUT = {
    '10MW': f'{ROOT}/pointmaps/models/emu/trina_10MW/emu.csv',
    '5.5MW': f'{ROOT}/pointmaps/models/emu/trina_5.5MW/emu.csv',
}

wb = openpyxl.load_workbook(XLSX, data_only=True)
pcs_ws = wb['PCS System(PCS 系统)']
sa_ws = wb['PCS Status&Alarm(PCS状态和报警)']

# ---------- Status&Alarm 位定义 ----------
RESERVED_NAMES = {'保留', 'reserve', 'resvd', '—', '-', ''}

def is_reserved(name):
    s = str(name).strip()
    return s in RESERVED_NAMES or '保留' in s or 'reserve' in s.lower() or 'resvd' in s.lower()

def cn_header(text):
    """提取表头括号内的中文名, 如 'System Fault Word1（系统故障字1）' -> '系统故障字1'"""
    m = re.findall(r'（([^（）]+)）', str(text))
    return m[-1].strip() if m else str(text).strip()

word_bits = {}  # 中文表头 -> [(bit, 名称)]
cur_left, cur_right = None, None
bit_re = re.compile(r'^Bit(\d+)$')
for r in range(1, sa_ws.max_row + 1):
    c1, c3 = sa_ws.cell(row=r, column=1).value, sa_ws.cell(row=r, column=3).value
    c6, c8 = sa_ws.cell(row=r, column=6).value, sa_ws.cell(row=r, column=8).value
    if c1 is not None and not bit_re.match(str(c1).strip()):
        cur_left = cn_header(c1)
        word_bits.setdefault(cur_left, [])
    if c6 is not None and not bit_re.match(str(c6).strip()):
        cur_right = cn_header(c6)
        word_bits.setdefault(cur_right, [])
    if c1 is not None:
        m = bit_re.match(str(c1).strip())
        if m and cur_left and c3 and not is_reserved(c3):
            word_bits[cur_left].append((int(m.group(1)), str(c3).strip()))
    if c6 is not None:
        m = bit_re.match(str(c6).strip())
        if m and cur_right and c8 and not is_reserved(c8):
            word_bits[cur_right].append((int(m.group(1)), str(c8).strip()))

for k in word_bits:
    word_bits[k].sort()

def bits_for(name):
    """寄存器名称 -> Status&Alarm 位定义; 先具体后一般"""
    rules = [
        ('中压综合故障字1', '中压综合故障字1'),
        ('PCS 中压状态字1', 'PCS中压状态字1'),
        ('PCS 中压状态字2', 'PCS中压状态字2'),
        ('系统高效模式1', '系统模式字1'),
        ('系统高效模式2', '系统模式字2'),
        ('系统状态字1', '系统状态字1'),
        ('系统状态字2', '系统状态字2'),
        ('系统状态字3', '系统状态字3'),
        ('系统故障字1', '系统故障字1'),
        ('系统故障字2', '系统故障字2'),
        ('系统故障字3', '系统故障字3'),
        ('系统故障字4', '系统故障字4'),
        ('系统警告字1', '系统警告字1'),
        ('系统警告字2', '系统警告字2'),
        ('控制器警告字1', '控制警告字1'),
        ('控制器故障字1', '控制器故障字1'),
        ('采样故障字1', '模块1/2 样本故障字1'),
        ('采样故障字2', '模块1/2 样本故障字2'),
        ('采样故障字3', '模块1/2 样本故障字3'),
        ('通信故障字1', '模块1/2 通信故障字1'),
        ('IO 故障字1', '模块1/2 IO报警字1'),
        ('IO 故障字2', '模块1/2 IO报警字2'),
        ('模块1 警告字1', '模块1/2 告警字1'),
        ('模块2 警告字1', '模块1/2 告警字1'),
        ('模块1 警告字2', '模块1/2 告警字2'),
        ('模块2 警告字2', '模块1/2 告警字2'),
    ]
    for i in range(1, 6):
        rules.insert(-6 - 2 * i, (f'FPGA 故障字{i}', f'FPAG故障字{i}'))
    for pat, key in rules:
        if pat in name:
            return word_bits.get(key, [])
    return []

# ---------- PCS System 寄存器解析 ----------
ADDR_RE = re.compile(r'^\s*(\d+)\s*\+\s*(\d+)\s*\*\s*\(n\s*-\s*1\)\s*\+\s*(\d+)\s*$')

SCALE_BY_UNIT = {
    '1': 1, '1V': 1, '1A': 1, '1kW': 1, '1kVar': 1, '1kVA': 1, 'kWh': 1, 'kVarh': 1,
    'min': 1, '%': 1, 'kW/s': 1, 'N/A': 1, '': 1,
    '0.1℃': 10, '0.1kV': 10, '0.1A': 10,
    '0.01Hz': 100, '0.01A': 100, '0.01Hz/s': 100, '0.01kW': 100, '0.01kVar': 100,
    '0.01V': 100, '0.01': 100,
    '0.001': 1000,
    '0.0001': 10000, '0.0001Hz': 10000,
}

def clean(s):
    if s is None:
        return ''
    s = re.sub(r'\s+', ' ', str(s)).strip()
    return s.replace(',', '，')

def parse_section(r1, r2):
    """解析一段寄存器定义, 返回 [(offset, 名称, type, unit)]"""
    regs = []
    for r in range(r1, r2 + 1):
        c3 = pcs_ws.cell(row=r, column=3).value   # 中文名
        c4 = pcs_ws.cell(row=r, column=4).value   # 地址公式
        c7 = pcs_ws.cell(row=r, column=7).value   # 单位
        c8 = pcs_ws.cell(row=r, column=8).value   # 数据类型
        if c4 is None:
            continue
        formula = str(c4)
        if '~' in formula or '\n' in formula:
            continue  # 保留区间
        m = ADDR_RE.match(formula)
        if not m:
            print(f'  ! 无法解析地址公式 R{r}: {formula!r}', file=sys.stderr)
            continue
        name = clean(c3)
        if not name or '保留' in name:
            continue
        if SKIP_WORD_RE.search(name):
            continue  # 按要求剔除的故障字寄存器
        if SKIP_INTEGRAL_RE.search(name):
            continue  # 剔除时间积分类字段（每小时/每日 容量、电量、循环时间、运行时间）
        regs.append((int(m.group(3)), name, clean(c8), clean(c7)))
    regs.sort(key=lambda x: x[0])
    return regs

# ---------- 寄存器段 ----------
# 剔除模块级 FPGA / IO / 采样 故障字寄存器 (仿真无对应故障源)
SKIP_WORD_RE = re.compile(r'模块[12]\s*(FPGA|IO|采样)\s*故障字')
# 剔除时间积分类字段: 每小时/每日的容量、电量、循环时间、运行时间 (高/低 16 位均命中)
SKIP_INTEGRAL_RE = re.compile(r'(每小时|每日).*(容量|电量|循环时间|运行时间)')

SECTIONS = {
    '10MW': {
        'unit_read':  parse_section(5, 237),    # 2600 + 300*(n-1), n=1..4
    },
    '5.5MW': {
        'unit_read':  parse_section(243, 461),  # 5000 + 600*(n-1), n=1..2
    },
}

LAYOUT = {
    '10MW': dict(base_r=2600, stride_r=300, n_r=4,
                 prefix=lambda n: f'10MW PCS{n}'),
    '5.5MW': dict(base_r=5000, stride_r=600, n_r=2,
                  prefix=lambda n: f'5.5MW PCS{n}'),
}

# ---------- 绑定规则 ----------
# 协议中“PCS”对应仿真中的 EMU 单元(emu{n}), 每台 EMU 含两台 PCS:
# 模块1 -> emu{n}.PcsList[0], 模块2 -> emu{n}.PcsList[1]
# 机绁级聚合优先 Emu.*, 无聚合属性的取 PcsList[0] 代表。
UNIT_BIND = {
    '交流电流 R': 'PcsList[0].PhaseACurrent',
    '交流电流 S': 'PcsList[0].PhaseBCurrent',
    '交流电流 T': 'PcsList[0].PhaseCCurrent',
    '电池总功率': 'PcsList[0].BatteryPower',
    '电池1 功率': 'PcsList[0].BatteryPower',
    '电池2 功率': 'PcsList[1].BatteryPower',
    '电网有功功率': 'Emu.OutputActivePower',
    '电网无功功率': 'Emu.OutputReactivePower',
    '电网电压 RS': 'PcsList[0].LineVoltageAB',
    '电网电压 ST': 'PcsList[0].LineVoltageBC',
    '电网电压 TR': 'PcsList[0].LineVoltageCA',
    'PCS 过温降载NTC': 'PcsList[0].IGBTMaxTemp',
    'PCS 可用容量': 'PcsList[0].AvailableCapacity',
    'PCS 额定容量': 'PcsList[0].PCSRatePower',
    '交流总充电有功电量低16位': 'PcsList[0].TotalChargeEnergy',
    '交流总放电有功电量低16位': 'PcsList[0].TotalDischargeEnergy',
}
MODULE_BIND = {
    '电池电压': 'BatteryVoltage',
    '电池电流': 'BatteryCurrent',
    '电感电流 R': 'PhaseACurrent',
    '电感电流 S': 'PhaseBCurrent',
    '电感电流 T': 'PhaseCCurrent',
    '总充电容量低16位': 'TotalChargeEnergy',
    '总放电容量低16位': 'TotalDischargeEnergy',
}

# 警告字插件组字: model=plugin|arg1=<字键>|arg2=<设备根路径>
# 位 -> 仿真故障映射由运行期 TrinaEmuFaultWordPlugin 维护, 不支持的位恒为 0
PLUGIN_BIND = {
    '模块1 警告字1': ('ModuleWarningWord1', 0),
    '模块1 警告字2': ('ModuleWarningWord2', 0),
    '模块2 警告字1': ('ModuleWarningWord1', 1),
    '模块2 警告字2': ('ModuleWarningWord2', 1),
}

def binding_for(name, n, scale):
    """按 EMU 层级返回 ModelSim 绑定; 无对应仿真属性返回 '0'。"""
    p = PLUGIN_BIND.get(name)
    if p:
        key, slot = p
        return f'model=plugin|arg1={key}|arg2=emu{n}.PcsList[{slot}]'
    path = None
    m = re.match(r'^模块([12])(.+)$', name)
    if m:
        prop = MODULE_BIND.get(m.group(2))
        if prop:
            path = f'PcsList[{int(m.group(1)) - 1}].{prop}'
    else:
        path = UNIT_BIND.get(name)
    if not path:
        return '0'
    return f'model=4|arg1=emu{n}.{path}|arg2=|arg3=|arg4={scale}'

# ---------- 生成 ----------
def csv_type(xl_type):
    return {'int16': 'int16', 'uint16': 'u16'}.get(xl_type, 'u16')

def scale_of(unit):
    s = SCALE_BY_UNIT.get(unit, 1)
    if unit and unit not in SCALE_BY_UNIT:
        print(f'  ! 未知单位 {unit!r}, 按 Scale=1 处理', file=sys.stderr)
    return s

def describe(name):
    bits = bits_for(name)
    if not bits:
        return name
    return name + '(' + ';'.join(f'Bit{b} {t}' for b, t in bits) + ')'

def build_rows(model):
    lay = LAYOUT[model]
    reads = []
    # PCS 单元读寄存器: 先按 offset 优先、跨 n 展开, 最后按地址升序输出
    for off, name, typ, unit in SECTIONS[model]['unit_read']:
        for n in range(1, lay['n_r'] + 1):
            addr = lay['base_r'] + lay['stride_r'] * (n - 1) + off
            reads.append((addr, f"{lay['prefix'](n)}-{describe(name)}", typ, unit, name, n))
    reads.sort(key=lambda x: x[0])
    rows = [['FunctionCode', 'Address', 'Type', 'Size', 'ParamName', 'Scale', 'Description', 'ModelSim']]
    for i, (addr, desc, typ, unit, name, n) in enumerate(reads):
        scale = scale_of(unit)
        rows.append(['4', str(addr), csv_type(typ), '16', f'yc{i}', str(scale), desc,
                     binding_for(name, n, scale)])
    return rows

for model, path in OUT.items():
    rows = build_rows(model)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerows(rows)
    n_read = sum(1 for r in rows[1:] if r[0] == '4')
    n_write = sum(1 for r in rows[1:] if r[0] == '6')
    n_bind = sum(1 for r in rows[1:] if r[7] != '0')
    print(f'{model}: 共 {len(rows) - 1} 点 (读 {n_read} / 写 {n_write}), 绑定 {n_bind} 点 -> {path}')

# Status&Alarm 覆盖情况自检
print('\nStatus&Alarm 位表:', {k: len(v) for k, v in word_bits.items() if v})
